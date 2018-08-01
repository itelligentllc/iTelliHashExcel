#! /usr/bin/env
# coding: utf-8
# itellihashexcel.py
# Copyright 2018 iTtelligent, LLC., Kirby J. Davis (kdavis@itelligentllc.com)

"""This file is part of iTelliHashExcel.

    iTelliHashExcel - A Cryptographic Hashing Application for Excel Files
    Copyright (C) 2018 iTtelligent, LLC (Kirby J. Davis)

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU Affero General Public License as published
    by the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU Affero General Public License for more details.

    You should have received a copy of the GNU Affero General Public License
    along with this program.  If not, see <http://www.gnu.org/licenses/>.
    """

import gettext
import os
import sys
import threading

import excelcryptohashinglogic as chl
import itellihashexcelimages_white as itellihashexcelimages
import wx
import wx.lib.scrolledpanel
from openpyxl import load_workbook
from wx.adv import AboutDialogInfo
from wx.lib.itemspicker import (ItemsPicker, EVT_IP_SELECTION_CHANGED, IP_SORT_CHOICES, IP_SORT_SELECTED)
from wx.lib.wordwrap import wordwrap

_licenseText = "iTelliHashExcel - A Cryptographic Hashing Application for Excel Files\n" \
               "Copyright (C) 2018  iTtelligent, LLC\n\n" \
               "This program is free software: you can redistribute it and/or modify it under the terms of the GNU " \
               "Affero General Public License as published by the Free Software Foundation, either version 3 of the " \
               "License, or (at your option) any later version.\n\n" \
               "This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without " \
               "even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU " \
               "Affero General Public License for more details.\n\n" \
               "You should have received a copy of the GNU Affero General Public License along with this program. " \
               "If not, see <http://www.gnu.org/licenses/>."

_utorange = wx.Colour(247, 130, 0)  # UT orange
_utsmokey = wx.Colour(88, 89, 91)  # Dark Gray. Use instead of black.
_utrock = wx.Colour(167, 169, 172)  # Light Gray
_utriver = wx.Colour(81, 124, 150)  # Gray-blue
_utlimestone = wx.Colour(240, 237, 227)  # Off white
_utlegacy = wx.Colour(87, 149, 132)  # Dark Green
_utswitchgrass = wx.Colour(171, 193, 120)  # Light Green
_utsunsphere = wx.Colour(254, 213, 53)  # Yellow
_utsummitt = wx.Colour(185, 225, 226)  # Light Blue-green
_utleconte = wx.Colour(141, 32, 72)  # Magenta
_uteureka = wx.Colour(235, 234, 100)  # Yellow-green
_utvalley = wx.Colour(0, 116, 111)  # Green
_utglobe = wx.Colour(0, 108, 147)  # Blue
_utregalia = wx.Colour(117, 74, 126)  # Purple
_utwhite = wx.Colour(255, 255, 255)  # White


class FieldsPickerDialog(wx.Dialog):
    """
    Present to user all fields available from the input file(s) selected in previous step that may
    be selected for hashing. Allow user to select desired fields/columns.
    """

    def __init__(self, parent, fieldsavailable):
        wx.Dialog.__init__(self, parent)
        self.fields2hash = []
        MainFrame.fields2hash = []
        choices = fieldsavailable.replace('\n', ',')
        choices = choices.replace('"', '')
        choices = choices.rstrip(',')
        choices = choices.split(',')
        choices = set(choices)
        choices = list(choices)
        sizer = wx.BoxSizer(wx.VERTICAL)
        self.ip = ItemsPicker(self, -1, choices, 'Available Columns:', 'Selected Columns:',
                              IP_SORT_CHOICES | IP_SORT_SELECTED)
        self.ip.Bind(EVT_IP_SELECTION_CHANGED, self.onselectionchange)
        self.ip.SetMinSize((-1, 150))
        sizer.Add(self.ip, 0, wx.ALL, 10)
        b = wx.Button(self, -1, "Click after finishing column selection(s)", style=wx.NO_BORDER)
        b.SetToolTip(
            'The information in the lists above should represent column names. If not, then your input data is not '
            'in the correct format. The selected sheet should contain the column names in the first row.')
        b.Bind(wx.EVT_BUTTON, self.onfinished)
        sizer.Add(b, 0, wx.ALL | wx.CENTER, 5)
        self.SetSizer(sizer)
        self.Fit()

    def onfinished(self, e):
        """ Once user clicks 'finished' button create list of fields to be hashed.

        :param e: Event
        :return: Fields to be hashed: fields2hash

        """
        if len(self.fields2hash) != 0:
            MainFrame.fields2hash = self.fields2hash.split(',')
        self.Close()

    def onselectionchange(self, e):
        """ Logic for identifying and selecting fields to be hashed.

        :param e: Event
        :return: Internal list of fields selected.
        """
        self.items = ",".join(e.GetItems())
        self.fields2hash = self.items


class WorkerThread(threading.Thread):
    """
    This section does the hashing logic routines within a thread separate from the main app GUI thread therefore
    allowing the GUI to remain responsive while the hashing logic is performed.

    """

    def __init__(self, threadNum, window):
        threading.Thread.__init__(self)
        self.threadNum = threadNum
        self.window = window
        self.timeToQuit = threading.Event()
        self.timeToQuit.clear()

    def stop(self):
        self.timeToQuit.set()

    def run(self):
        wx.CallAfter(self.window.statusBar.SetLabel, "Creating temporary database... please wait...")
        mychl.create_temp_db(self.window.fileselected, self.window.sheet2process, self.window.fields2hash,
                             self.window.cols2hash, self.window.inputdirectory)
        wx.CallAfter(self.window.statusBar.SetLabel, "Creating & writing summary mapping file... please wait...")
        mychl.process_hash_mapfile_summary(self.window.fileextension, self.window.outputdirectory)
        wx.CallAfter(self.window.statusBar.SetLabel, "Creating & writing detail mapping file... please wait...")
        mychl.process_hash_mapfile_detail(self.window.fields2hash, self.window.fileextension,
                                          self.window.outputdirectory)
        wx.CallAfter(self.window.statusBar.SetLabel,
                     "Creating & writing output file with a separate sheet for each selected column... please wait...")
        mychl.create_hashed_outputfile(self.window.fileselected, self.window.sheet2process, self.window.fileextension,
                                       self.window.inputdirectory,
                                       self.window.outputdirectory)
        mychl.remove_sqlite()
        self.timeToQuit.set()
        self.window.onlongrundone()


class MainFrame(wx.Frame):
    """ Main frame of Excel Cryptographic Hashing program

    """

    def __init__(self):
        wx.Frame.__init__(self, None, id=wx.ID_ANY,
                          title="iTelliHashExcel - A Cryptographic Hashing Application for Excel Files",
                          pos=wx.DefaultPosition,
                          size=wx.Size(600, 465), style=wx.DEFAULT_FRAME_STYLE | wx.TAB_TRAVERSAL)
        # set threading
        self.threads = []
        self.count = 0

        # set window icon
        self.icon = itellihashexcelimages.MyIcon.GetIcon()
        self.logo = itellihashexcelimages.MyLogo.GetBitmap()
        self.SetIcon(self.icon)

        # set colors for UI
        self.SetBackgroundColour(_utlimestone)
        self.selectable = _utorange
        self.unselectable = _utlimestone
        self.msgbar = _utswitchgrass

        bSizer_Frame_V = wx.BoxSizer(wx.VERTICAL)

        bSizer_Step1_V = wx.BoxSizer(wx.VERTICAL)

        bSizer_Bitmap = wx.BoxSizer(wx.HORIZONTAL)

        bSizer_Step1_V.Add(bSizer_Bitmap, 1, wx.EXPAND, 5)
        self.bitmap1 = wx.StaticBitmap(self, wx.ID_ANY,
                                       self.logo,
                                       wx.DefaultPosition, wx.DefaultSize, 0)
        bSizer_Step1_V.Add(self.bitmap1, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL | wx.EXPAND, 5)

        # Step 1
        self.static_Step1 = wx.StaticText(self, wx.ID_ANY, "Step 1: \nSelect Desired Cryptographic Hash Algorithm ",
                                          wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTRE)
        self.static_Step1.Wrap(-1)
        self.static_Step1.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        bSizer_Step1_V.Add(self.static_Step1, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 5)
        self.staticline_Step1_Top = wx.StaticLine(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                                  wx.LI_HORIZONTAL)
        bSizer_Step1_V.Add(self.staticline_Step1_Top, 0, wx.EXPAND | wx.ALL, 5)
        gSizer_Step1 = wx.GridSizer(1, 0, 0, 0)

        self.radioBtn_None = wx.RadioButton(self, wx.ID_ANY, "None", wx.DefaultPosition, wx.DefaultSize, 0)
        self.radioBtn_None.SetToolTip("Select one of the cryptographic hash algorithms to begin processing. "
                                      "Please note that as the bit size increases, both the time to process "
                                      "the input file(s) and the size of your output data files will increase.")
        gSizer_Step1.Add(self.radioBtn_None, 0, wx.ALIGN_CENTER | wx.ALL, 5)

        self.radioBtn_RipeMD = wx.RadioButton(self, wx.ID_ANY, "RIPEMD-160", wx.DefaultPosition, wx.DefaultSize, 0)
        self.radioBtn_RipeMD.SetToolTip(
            "RIPEMD-160 is an improved, 160-bit version of the original RIPEMD, and the most common version in the "
            "family. RIPEMD-160 was designed in the open academic community, in contrast to the NSA designed "
            "SHA-1 and SHA-2 algorithms.")
        gSizer_Step1.Add(self.radioBtn_RipeMD, 0, wx.ALIGN_CENTER | wx.ALL, 5)

        self.radioBtn_SHA224 = wx.RadioButton(self, wx.ID_ANY, "SHA-224", wx.DefaultPosition, wx.DefaultSize, 0)
        self.radioBtn_SHA224.SetToolTip(
            "A 224-bit hash function within the SHA-2 (Secure Hash Algorithm 2) family. SHA-224 is an approved "
            "secure hash standard (SHS) per U.S. National Institute of Standards and Technology (NIST) FIPS PUB "
            "140-2, Annex A.")
        gSizer_Step1.Add(self.radioBtn_SHA224, 0, wx.ALIGN_CENTER | wx.ALL, 5)

        self.radioBtn_SHA256 = wx.RadioButton(self, wx.ID_ANY, "SHA-256", wx.DefaultPosition, wx.DefaultSize, 0)
        self.radioBtn_SHA256.SetToolTip(
            "A 256-bit hash function within the SHA-2 (Secure Hash Algorithm 2) family. SHA-256 is an approved "
            "secure hash standard (SHS) per U.S. National Institute of Standards and Technology (NIST) FIPS PUB "
            "140-2, Annex A.")
        gSizer_Step1.Add(self.radioBtn_SHA256, 0, wx.ALIGN_CENTER | wx.ALL, 5)

        self.radioBtn_SHA384 = wx.RadioButton(self, wx.ID_ANY, "SHA-384", wx.DefaultPosition, wx.DefaultSize, 0)
        self.radioBtn_SHA384.SetToolTip(
            "A 384-bit hash function within the SHA-2 (Secure Hash Algorithm 2) family. SHA-384 is an approved "
            "secure hash standard (SHS) per U.S. National Institute of Standards and Technology (NIST) FIPS PUB "
            "140-2, Annex A.")
        gSizer_Step1.Add(self.radioBtn_SHA384, 0, wx.ALIGN_CENTER | wx.ALL, 5)

        self.radioBtn_SHA512 = wx.RadioButton(self, wx.ID_ANY, "SHA-512", wx.DefaultPosition, wx.DefaultSize, 0)
        self.radioBtn_SHA512.SetToolTip(
            "A 512-bit hash function within the SHA-2 (Secure Hash Algorithm 2) family. SHA-512 is an approved "
            "secure hash standard (SHS) per U.S. National Institute of Standards and Technology (NIST) FIPS PUB "
            "140-2, Annex A.")
        gSizer_Step1.Add(self.radioBtn_SHA512, 0, wx.ALIGN_CENTER | wx.ALL, 5)

        bSizer_Step1_V.Add(gSizer_Step1, 1, wx.ALIGN_CENTER_HORIZONTAL | wx.EXPAND, 5)
        self.staticline_Step1_Bottom = wx.StaticLine(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                                     wx.LI_HORIZONTAL)
        bSizer_Step1_V.Add(self.staticline_Step1_Bottom, 0, wx.EXPAND | wx.ALL, 5)
        bSizer_Frame_V.Add(bSizer_Step1_V, 0, wx.ALL | wx.EXPAND, 5)

        # Step 2
        gbSizer_Step2_4 = wx.GridBagSizer(0, 0)
        gbSizer_Step2_4.SetFlexibleDirection(wx.BOTH)
        gbSizer_Step2_4.SetNonFlexibleGrowMode(wx.FLEX_GROWMODE_SPECIFIED)
        self.staticText_Step2 = wx.StaticText(self, wx.ID_ANY, "Step 2:", wx.DefaultPosition, wx.DefaultSize, 0)
        self.staticText_Step2.Wrap(-1)
        self.staticText_Step2.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        gbSizer_Step2_4.Add(self.staticText_Step2, wx.GBPosition(0, 0), wx.GBSpan(1, 1), wx.ALL, 10)
        self.button_Step2 = wx.Button(self, wx.ID_ANY, "Select Excel File and Sheet to Process", wx.DefaultPosition,
                                      wx.DefaultSize, style=wx.NO_BORDER)

        self.button_Step2.SetToolTip(
            'You may only select one file and sheet to process at one time. The Excel file should contain the column'
            ' names in the first row.')
        self.button_Step2.Enable(False)
        gbSizer_Step2_4.Add(self.button_Step2, wx.GBPosition(0, 1), wx.GBSpan(1, 4), wx.ALL | wx.EXPAND, 5)

        # Step 3
        self.staticText_Step3 = wx.StaticText(self, wx.ID_ANY, "Step 3:", wx.DefaultPosition, wx.DefaultSize, 0)
        self.staticText_Step3.Wrap(-1)
        self.staticText_Step3.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        gbSizer_Step2_4.Add(self.staticText_Step3, wx.GBPosition(1, 0), wx.GBSpan(1, 1), wx.ALL, 10)
        self.button_Step3 = wx.Button(self, wx.ID_ANY, "Select Column(s) to Cryptographically Hash",
                                      wx.DefaultPosition, wx.DefaultSize, style=wx.NO_BORDER)
        self.button_Step3.SetToolTip(
            "In the following window, please select the column(s) you wish to hash. You may select as many columns as "
            "desired. However, processing time and file size will increase as the number of columns selected increase.")
        self.button_Step3.Enable(False)
        gbSizer_Step2_4.Add(self.button_Step3, wx.GBPosition(1, 1), wx.GBSpan(1, 4), wx.ALL | wx.EXPAND, 5)

        # Step 4
        self.staticText_Step4A = wx.StaticText(self, wx.ID_ANY, "Step 4:", wx.DefaultPosition, wx.DefaultSize, 0)
        self.staticText_Step4A.Wrap(-1)
        self.staticText_Step4A.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 92, False, wx.EmptyString))
        gbSizer_Step2_4.Add(self.staticText_Step4A, wx.GBPosition(2, 0), wx.GBSpan(1, 1), wx.ALL, 10)
        # Step 4A
        self.button_Step4A = wx.Button(self, wx.ID_ANY, "Start Hashing (Results to Input Folder)", wx.DefaultPosition,
                                       wx.DefaultSize,
                                       style=wx.NO_BORDER)
        self.button_Step4A.SetToolTip(
            "Pressing this button will start the hashing process and save the output files to the same folder as "
            "the Excel input file.")
        self.button_Step4A.Enable(False)
        gbSizer_Step2_4.Add(self.button_Step4A, wx.GBPosition(2, 1), wx.GBSpan(1, 2), wx.ALL | wx.EXPAND, 5)
        # Step 4B
        self.button_Step4B = wx.Button(self, wx.ID_ANY, "Start Hashing (Results to a Different Folder)",
                                       wx.DefaultPosition,
                                       wx.DefaultSize,
                                       style=wx.NO_BORDER)
        self.button_Step4B.SetToolTip(
            "Pressing this button will start the hashing process after selecting or creating a new folder location "
            "for saving the output files.")
        self.button_Step4B.Enable(False)
        gbSizer_Step2_4.Add(self.button_Step4B, wx.GBPosition(2, 3), wx.GBSpan(1, 2), wx.ALL | wx.EXPAND, 5)

        # Progress Gauge
        self.gauge_progress = wx.Gauge(self, wx.ID_ANY, 100, wx.DefaultPosition, wx.DefaultSize, wx.GA_HORIZONTAL)
        gbSizer_Step2_4.Add(self.gauge_progress, wx.GBPosition(3, 1), wx.GBSpan(1, 4), wx.ALL | wx.EXPAND, 10)
        bSizer_Frame_V.Add(gbSizer_Step2_4, 1, wx.ALL | wx.EXPAND, 5)

        # Info and Close Buttons
        gSizer2 = wx.GridSizer(0, 2, 0, 0)
        self.button_Info = wx.Button(self, wx.ID_ANY, "Information", wx.DefaultPosition, wx.DefaultSize,
                                     wx.NO_BORDER)
        gSizer2.Add(self.button_Info, 0, wx.ALL | wx.EXPAND, 5)

        self.button_Close = wx.Button(self, wx.ID_ANY, "Exit", wx.DefaultPosition, wx.DefaultSize, wx.NO_BORDER)
        gSizer2.Add(self.button_Close, 0, wx.ALL | wx.EXPAND, 5)
        bSizer_Frame_V.Add(gSizer2, 0, wx.ALL | wx.EXPAND, 5)

        # Status Bar
        self.SetSizer(bSizer_Frame_V)
        self.Layout()
        self.statusBar = self.CreateStatusBar(1, wx.STB_SIZEGRIP, wx.ID_ANY)
        self.statusBar.SetBackgroundColour(self.msgbar)

        self.Centre(wx.BOTH)

        # Connect Events
        self.radioBtn_None.Bind(wx.EVT_RADIOBUTTON, self.radioBtn_NoneOnRadioButton)
        self.radioBtn_RipeMD.Bind(wx.EVT_RADIOBUTTON, self.radioBtn_RipeMDOnRadioButton)
        self.radioBtn_SHA224.Bind(wx.EVT_RADIOBUTTON, self.radioBtn_SHA224OnRadioButton)
        self.radioBtn_SHA256.Bind(wx.EVT_RADIOBUTTON, self.radioBtn_SHA256OnRadioButton)
        self.radioBtn_SHA384.Bind(wx.EVT_RADIOBUTTON, self.radioBtn_SHA384OnRadioButton)
        self.radioBtn_SHA512.Bind(wx.EVT_RADIOBUTTON, self.radioBtn_SHA512OnRadioButton)
        self.button_Step2.Bind(wx.EVT_BUTTON, self.button_Step2OnButtonClick)
        self.button_Step3.Bind(wx.EVT_BUTTON, self.button_Step3OnButtonClick)
        self.button_Step4A.Bind(wx.EVT_BUTTON, self.button_Step4AOnButtonClick)
        self.button_Step4B.Bind(wx.EVT_BUTTON, self.button_Step4BOnButtonClick)
        self.button_Info.Bind(wx.EVT_BUTTON, self.button_InfoOnButtonClick)
        self.button_Close.Bind(wx.EVT_BUTTON, self.button_CloseOnButtonClick)

        self.Show()

    @property
    def fields2hash(self):
        return self.fields2hash

    def onlongrundone(self):
        self.gauge_progress.SetValue(100)
        self.hash2use = 0
        self.radioBtn_RipeMD.Enable(True)
        self.radioBtn_SHA224.Enable(True)
        self.radioBtn_SHA256.Enable(True)
        self.radioBtn_SHA384.Enable(True)
        self.radioBtn_SHA512.Enable(True)
        self.radioBtn_None.Enable(True)
        self.radioBtn_None.SetValue(1)
        self.button_Step2.Enable(False)
        self.button_Step2.SetBackgroundColour(self.unselectable)
        self.button_Step3.SetBackgroundColour(self.unselectable)
        self.button_Step4A.SetBackgroundColour(self.unselectable)
        self.button_Step4B.SetBackgroundColour(self.unselectable)
        self.statusBar.SetLabel("Finished !! You may now exit or process another input file.")

    def radioBtn_NoneOnRadioButton(self, event):
        """ STEP 1. 'None' Hash format selection button. This button is initially 'selected' when the
        program is started. User must select from among the hashing options available to begin
        hashing process through successive steps via input buttons and dialogs.

        :param event: Event
        :return: hash2use value for subsequent processing. Not used in further processing.

        """
        self.hash2use = 0
        self.button_Step2.Enable(False)
        self.statusBar.SetLabel("Step 1: Please select Cryptographic Hashing Algorithm")
        self.button_Step2.SetBackgroundColour(self.unselectable)
        self.button_Step3.SetBackgroundColour(self.unselectable)
        self.button_Step4A.SetBackgroundColour(self.unselectable)
        self.button_Step4B.SetBackgroundColour(self.unselectable)

    def radioBtn_RipeMDOnRadioButton(self, event):
        """ STEP 1. 'RIPEMD-160' Hash format selection button

        :param event: Event
        :return: hash2use value for subsequent processing.

        """
        self.hash2use = 1
        self.button_Step2.Show()
        self.button_Step2.Enable(True)
        self.statusBar.SetLabel("RIPEMD-160 Cryptographic Hash has been selected. Ready for Step 2.")
        self.button_Step2.SetBackgroundColour(self.selectable)

    def radioBtn_SHA224OnRadioButton(self, event):
        """ STEP 1. 'SHA-224' Hash format selection button

        :param event: Event
        :return: hash2use value for subsequent processing.

        """
        self.hash2use = 2
        self.button_Step2.Show()
        self.button_Step2.Enable(True)
        self.statusBar.SetLabel("SHA-224 Cryptographic Hash has been selected. Ready for Step 2.")
        self.button_Step2.SetBackgroundColour(self.selectable)

    def radioBtn_SHA256OnRadioButton(self, event):
        """ STEP 1. 'SHA-256' Hash format selection button

        :param event: Event
        :return: hash2use value for subsequent processing.

        """
        self.hash2use = 3
        self.button_Step2.Show()
        self.button_Step2.Enable(True)
        self.statusBar.SetLabel("SHA-256 Cryptographic Hash has been selected. Ready for Step 2.")
        self.button_Step2.SetBackgroundColour(self.selectable)

    def radioBtn_SHA384OnRadioButton(self, event):
        """ STEP 1. 'SHA-384' Hash format selection button

        :param event: Event
        :return: hash2use value for subsequent processing.

        """
        self.hash2use = 4
        self.button_Step2.Show()
        self.button_Step2.Enable(True)
        self.statusBar.SetLabel("SHA-384 Cryptographic Hash has been selected. Ready for Step 2.")
        self.button_Step2.SetBackgroundColour(self.selectable)

    def radioBtn_SHA512OnRadioButton(self, event):
        """ STEP 1. 'SHA-512' Hash format selection button

        :param event: Event
        :return: hash2use value for subsequent processing.

        """
        self.hash2use = 5
        self.button_Step2.Show()
        self.button_Step2.Enable(True)
        self.statusBar.SetLabel("SHA-512 Cryptographic Hash has been selected. Ready for Step 2.")
        self.button_Step2.SetBackgroundColour(self.selectable)

    def button_Step2OnButtonClick(self, event):
        """ STEP 2. Present a wxpython FileDialog to allow user to select Excel input file(s) for processing.
        Files with suffixes ending in .xls, .xlsx, or any (*.*) may be selected.

        :param event: Event
        :return: fileselected: List of Excel input files selected for processsing
                 fieldsavailable: List of available fields/columns from the Excel input file
                                  and sheet selected that are available for processing.

        """
        self.radioBtn_None.Enable(False)
        self.radioBtn_RipeMD.Enable(False)
        self.radioBtn_SHA224.Enable(False)
        self.radioBtn_SHA256.Enable(False)
        self.radioBtn_SHA384.Enable(False)
        self.radioBtn_SHA512.Enable(False)
        wildcard = "Excel 2007+ files (*.xlsx;*.xlsm)|*.xlsx;*.xlsm"
        dialog1A = wx.FileDialog(self,
                                 message="Choose an Excel file",
                                 defaultDir=os.path.expanduser("~"),
                                 defaultFile="",
                                 wildcard=wildcard,
                                 style=wx.FD_OPEN | wx.FD_CHANGE_DIR
                                 )
        if dialog1A.ShowModal() == wx.ID_OK:
            self.statusBar.SetLabel("Please wait... reading and loading Excel input file.")
            self.inputdirectory = dialog1A.GetDirectory() + '\\'
            self.outputdirectory = self.inputdirectory
            self.fileselected = dialog1A.GetFilename()
            self.fileextension = os.path.splitext(dialog1A.GetPath())[1]
            self.sheetsavailable = load_workbook(filename=self.fileselected, read_only=True,
                                                 keep_vba=False).get_sheet_names()
            dialog1B = wx.SingleChoiceDialog(
                self, 'Please select sheet to process', 'Sheet Selection',
                self.sheetsavailable,
                wx.CHOICEDLG_STYLE
            )
            if dialog1B.ShowModal() == wx.ID_OK:
                self.sheet2process = dialog1B.GetStringSelection()
                dialog1B.Destroy()
                data = load_workbook(filename=self.fileselected, read_only=True, keep_vba=False)
                sheet = data.get_sheet_by_name(self.sheet2process)
                try:
                    self.myDict = {}
                    for i in range(1, sheet.max_column + 1, 1):
                        self.myDict[sheet.cell(row=1, column=i).value] = i - 1
                    self.fieldsavailable = ",".join(list(self.myDict.keys()))
                    self.button_Step2.Enable(False)
                    self.button_Step2.SetBackgroundColour(self.unselectable)
                    self.button_Step3.Enable(True)
                    self.button_Step3.SetBackgroundColour(self.selectable)
                    self.statusBar.SetLabel("Excel input file has been selected and loaded. Ready for Step 3.")
                    dialog1A.Destroy()
                except:
                    self.statusBar.SetLabel(
                        "Error: Selected sheet contains invalid or no data. Please select another file or sheet.")

    def button_Step3OnButtonClick(self, event):
        """ STEP 3. (See FieldsPickerDialog) Present to user all fields available from the input file(s) selected
        in previous step that may be selected for hashing. Allow user to select desired fields/columns.

        :param event: Event
        :return: fields2hash - List of fields selected by user to be hashed.

        """
        step3 = FieldsPickerDialog(self, self.fieldsavailable)
        step3.ShowModal()
        if len(self.fields2hash) != 0:
            MainFrame.cols2hash = list(self.myDict.get(i) for i in self.fields2hash)
            self.button_Step3.Enable(False)
            self.button_Step3.SetBackgroundColour(self.unselectable)
            self.button_Step4A.Enable(True)
            self.button_Step4A.SetBackgroundColour(self.selectable)
            self.button_Step4B.Enable(True)
            self.button_Step4B.SetBackgroundColour(self.selectable)
            self.statusBar.SetLabel("Column(s) have been selected for hashing. Ready for Step 4.")
        elif len(self.fields2hash) == 0:
            self.button_Step3.Enable(True)
            self.button_Step3.SetBackgroundColour(self.selectable)
            self.statusBar.SetLabel("Please select column(s) for hashing.")

    def button_Step4AOnButtonClick(self, event):
        """ STEP 4. Begin the hashing process. First identify/set the hashing format chosen and then
        begin the hashing process. Present user with "wait" message and start the progress bar.

        :param event: Event
        :return: Excel files written to disk at same location as Excel input files selected for processing.

        """
        self.button_Step4A.Enable(False)
        self.button_Step4A.SetBackgroundColour(self.unselectable)
        self.button_Step4B.Enable(False)
        self.button_Step4B.SetBackgroundColour(self.unselectable)
        mychl.initialize_sqlite()
        mychl.identify_hash(self.hash2use)
        self.gauge_progress.Pulse()
        self.statusBar.SetLabel("Setting up processing thread... please wait...")
        try:
            self.count += 1
            thread = WorkerThread(self.count, self)
            self.threads.append(thread)
            thread.daemon = True
            thread.start()
        except:
            self.statusBar.SetLabel("Unable to start processing thread")

    def button_Step4BOnButtonClick(self, event):
        """ STEP 4. Begin the hashing process. First identify/set the hashing format chosen and then
        begin the hashing process. Present user with "wait" message and start the progress bar.

        :param event: Event
        :return: Excel files written to disk at same location as Excel input files selected for processing.

        """
        self.button_Step4A.Enable(False)
        self.button_Step4A.SetBackgroundColour(self.unselectable)
        self.button_Step4B.Enable(False)
        self.button_Step4B.SetBackgroundColour(self.unselectable)
        dialog2 = wx.DirDialog(self, "Choose or create a folder for the output files:",
                               style=wx.DD_DEFAULT_STYLE | wx.DD_CHANGE_DIR)
        if dialog2.ShowModal() == wx.ID_OK:
            self.outputdirectory = dialog2.GetPath() + '\\'
        dialog2.Destroy()
        mychl.initialize_sqlite()
        mychl.identify_hash(self.hash2use)
        self.gauge_progress.Pulse()
        self.statusBar.SetLabel("Setting up processing thread... please wait...")
        try:
            self.count += 1
            thread = WorkerThread(self.count, self)
            self.threads.append(thread)
            thread.daemon = True
            thread.start()
        except:
            self.statusBar.SetLabel("Unable to start processing thread")

    def button_CloseOnButtonClick(self, event):
        self.Destroy()

    def button_InfoOnButtonClick(self, event):
        # First we create and fill the info object
        info = wx.adv.AboutDialogInfo()
        info.SetIcon(self.icon)
        info.SetName("iTelliHashExcel - A Cryptographic Hashing Application for Excel Files")
        info.SetVersion("1.0.1")
        info.SetCopyright("(C) 2018 iTelligent, LLC")
        info.SetDescription(wordwrap(
            "Please contact itellihashexcel@itelligentllc.com to report any bugs or to inquire about the availability "
            "of other licenses, such as traditional commercial licenses. Additionally, contact us to inquire about "
            "any additional services, such as product customization for non-standard Excel files, input file validation "
            "and correction, etc. At no time will iTelligent, LLC share any of your data with others. \n\n"
            "Please note: Since this program is intended to protect sensitive information through cryptographic "
            "hashing, it DOES NOT communicate in any way with any programs on your computer (other than Microsoft "
            "Excel) or with any other external programs, computers, or sites.\n\n"
            "Prerequisites: [1] Microsoft Excel 2007 or above is installed. [2] An Excel file (*.xlsx or *.xlsm) "
            "containing column names in the first row to process.\n\n"
            "This program will create three types of files in either the same directory as the selected Excel "
            "input file or a directory of the user's choosing:\n\n"
            "(1) A summary 'mapping' file containing the hashes and original/plaintext forms of all columns "
            "selected to be hashed. This file will be named 'Hash_MapFile_Summary_<selected hash format>'. This "
            "file will contain the column name (column name = ColumnName), original (column name = Plaintext), and "
            "hash (column name = HashValue) of each distinct value of all hashed columns in the original Excel file.\n\n"
            "(2) A detail 'mapping' file containing a sheet for every column hashed. This file will be named "
            "'Hash_MapFile_Detail_<selected hash format>'. Each sheet will contain the original/plaintext "
            "(column name = Plaintext) and hash (column name = Hashvalue) of each distinct value found in the column.\n\n"
            "(3) The original Excel file with 'mapping' sheet(s) added for each column hashed. The file will be "
            "named 'Hashed_<Original Excel File Name>_<selected hash format>'.\n\n"
            "Please note that due to the way Excel stores data, the original/plaintext values in the files produced "
            "may not match the original input values. For example, if one of your columns selected to be hashed "
            "contained social security numbers beginning with zeros, these 'leading' zeroes would not appear in the "
            "original/plaintext version. Please be aware of such possibilities in your data and adjust your "
            "processing accordingly.\n\n"
            "Please see iTelliHashExcel's product page for further information and examples as to how this program "
            "may be used to protect your sensitive information yet allow you to share your data and enlist the "
            "aid of other organizations and services.",
            800, wx.ClientDC(self)))
        info.SetWebSite("http://www.itelligentllc.com/itellihashexcel", "iTelliHashExcel Product Page")
        info.AddDeveloper("Kirby J. Davis")

        info.SetLicense(wordwrap(_licenseText, 600, wx.ClientDC(self)))

        wx.adv.AboutBox(info)

    @fields2hash.setter
    def fields2hash(self, value):
        self._fields2hash = value

    @fields2hash.setter
    def fields2hash(self, value):
        self._fields2hash = value


if __name__ == "__main__":

    try:
        app = wx.App(False)

        # Get the locale directory
        basepath = os.path.abspath(os.path.dirname(sys.argv[0]))
        localedir = os.path.join(basepath, "locale")
        langid = wx.LANGUAGE_DEFAULT  # use OS default; or use LANGUAGE_JAPANESE, etc.
        domain = "messages"  # the translation file is messages.mo
        # Set locale for wxWidgets
        mylocale = wx.Locale(langid)
        mylocale.AddCatalogLookupPathPrefix(localedir)
        mylocale.AddCatalog(domain)

        # Set up Python's gettext
        mytranslation = gettext.translation(domain, localedir, [mylocale.GetCanonicalName()], fallback=True)
        mytranslation.install()

        mychl = chl.ExcelCryptoHash()
        frame = MainFrame()
        app.MainLoop()
    except:
        import sys, traceback

        xc = traceback.format_exception(*sys.exc_info())
        wx.MessageBox(''.join(xc))
